from selenium import webdriver
from selenium.webdriver.edge.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException, NoSuchElementException, WebDriverException
import time
from datetime import datetime
import os
from PIL import Image, ImageDraw
from openpyxl import load_workbook
import logging
import traceback
import sys

# 配置日志
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler('网页截图_debug.log', encoding='utf-8'),
        logging.StreamHandler()
    ]
)
logger = logging.getLogger(__name__)

def create_date_folder(date_str):
    """创建以日期命名的文件夹"""
    # 提取月日部分，例如20250828 -> 08-28
    month = date_str[4:6]  # 取MM部分
    day = date_str[6:8]    # 取DD部分
    folder_name = f"{month}-{day}"  # 格式化为MM-DD
    folder_path = os.path.join(os.getcwd(), folder_name)
    
    if not os.path.exists(folder_path):
        os.makedirs(folder_path)
        logger.info(f"创建文件夹: {folder_path}")
    else:
        logger.info(f"文件夹已存在: {folder_path}")
    # 新增：复制当前目录下的“日常检查表.xlsx”到日期文件夹
    try:
        src_excel = os.path.join(os.getcwd(), "日常检查表.xlsx")
        dst_excel = os.path.join(folder_path, "日常检查表.xlsx")
        if os.path.exists(src_excel):
            if not os.path.exists(dst_excel):
                import shutil
                shutil.copy2(src_excel, dst_excel)
                logger.info(f"已复制Excel到日期文件夹: {dst_excel}")
            else:
                logger.info(f"日期文件夹中Excel已存在，跳过复制: {dst_excel}")
        else:
            logger.warning(f"源Excel不存在: {src_excel}")
    except Exception as copy_e:
        logger.error(f"复制Excel到日期文件夹失败: {str(copy_e)}")
        logger.error(f"详细错误信息: {traceback.format_exc()}")    
    return folder_path

def extract_svg_data_for_j_column(driver, page_num):
    """从第一个和第二个SVG中提取特定class的文本内容，用于J列"""
    logger.info(f"开始从第{page_num}个页面提取SVG数据用于J列")
    
    cpan1 = None
    dpan1 = None
    
    try:
        # 等待页面加载
        time.sleep(2)
        
        # 查找所有SVG元素
        svg_elements = driver.find_elements(By.TAG_NAME, "svg")
        logger.info(f"找到 {len(svg_elements)} 个SVG元素")
        
        # 从第一个SVG中提取数据
        if len(svg_elements) >= 1:
            svg_element = svg_elements[0]  # 第一个SVG
            logger.info("正在查找第一个SVG中的特定class元素")
            
            try:
                target_elements = svg_element.find_elements(By.CLASS_NAME, "svg-gauge-value-and-units")
                logger.info(f"在第一个SVG中找到 {len(target_elements)} 个svg-gauge-value-and-units元素")
                
                for element in target_elements:
                    try:
                        # 使用更安全的方式获取class属性
                        class_attr = driver.execute_script("return arguments[0].className.baseVal || arguments[0].className;", element)
                        if "svg-gauge-value-and-units-horizontal" in str(class_attr):
                            text_content = element.text.strip()
                            cpan1 = f"C:\\{text_content}"
                            logger.info(f"从第一个SVG提取到数据: {cpan1}")
                            break
                    except Exception as attr_e:
                        logger.warning(f"获取第一个SVG元素属性时出错: {str(attr_e)}")
                        continue
                        
            except Exception as e:
                logger.error(f"处理第一个SVG时出错: {str(e)}")
        
        # 从第二个SVG中提取数据
        if len(svg_elements) >= 2:
            svg_element = svg_elements[1]  # 第二个SVG
            logger.info("正在查找第二个SVG中的特定class元素")
            
            try:
                target_elements = svg_element.find_elements(By.CLASS_NAME, "svg-gauge-value-and-units")
                logger.info(f"在第二个SVG中找到 {len(target_elements)} 个svg-gauge-value-and-units元素")
                
                for element in target_elements:
                    try:
                        # 使用更安全的方式获取class属性
                        class_attr = driver.execute_script("return arguments[0].className.baseVal || arguments[0].className;", element)
                        if "svg-gauge-value-and-units-horizontal" in str(class_attr):
                            text_content = element.text.strip()
                            dpan1 = f"D:\\{text_content}"
                            logger.info(f"从第二个SVG提取到数据: {dpan1}")
                            break
                    except Exception as attr_e:
                        logger.warning(f"获取第二个SVG元素属性时出错: {str(attr_e)}")
                        continue
                        
            except Exception as e:
                logger.error(f"处理第二个SVG时出错: {str(e)}")
        
        # 如果没有提取到数据，设置默认值
        if not cpan1:
            cpan1 = f"C:\\页面{page_num}SVG1数据提取失败"
            logger.warning(f"第{page_num}个页面第一个SVG未能提取到数据")
            
        if not dpan1:
            dpan1 = f"D:\\页面{page_num}SVG2数据提取失败"
            logger.warning(f"第{page_num}个页面第二个SVG未能提取到数据")
        
        # 组合数据
        combined_data = f"{cpan1}\n{dpan1}"
        logger.info(f"组合后的数据: {combined_data}")
        
        return combined_data
        
    except Exception as e:
        logger.error(f"提取SVG数据时发生错误: {str(e)}")
        return f"C:\\提取失败\nD:\\提取失败"

def extract_data_to_excel(driver, cell_address_h, cell_address_i, cell_address_j, page_num, date_folder):
    # 提取网页数据并保存到Excel文件
    logger.info(f"开始提取第{page_num}个页面的数据到单元格 {cell_address_h}, {cell_address_i}, {cell_address_j}")
    
    try:
        # 记录当前页面URL
        current_url = driver.current_url
        logger.info(f"当前页面URL: {current_url}")
        
        # 等待页面加载
        logger.info("等待页面加载完成...")
        time.sleep(2)
        
        # 检查页面是否加载完成
        try:
            WebDriverWait(driver, 10).until(
                lambda d: d.execute_script("return document.readyState") == "complete"
            )
            logger.info("页面加载状态: 完成")
        except TimeoutException:
            logger.warning("页面加载超时，但继续尝试提取数据")
        
        # 记录页面标题
        page_title = driver.title
        logger.info(f"页面标题: {page_title}")
        
        data_value_svg3 = None
        data_value_svg5 = None
        
        try:
            # 查找SVG相关元素
            logger.info("查找SVG相关元素")
            svg_elements = driver.find_elements(By.TAG_NAME, "svg")
            logger.info(f"找到 {len(svg_elements)} 个SVG元素")
            
            # 查找第3个SVG元素
            if len(svg_elements) >= 3:
                svg_element = svg_elements[2]  # 第3个SVG（索引为2）
                logger.info("正在查找第3个SVG中的特定class元素")
                
                try:
                    # 在第3个SVG内查找class为svg-gauge-value-and-units svg-gauge-value-and-units-horizontal的元素
                    target_elements = svg_element.find_elements(By.CLASS_NAME, "svg-gauge-value-and-units")
                    logger.info(f"在第3个SVG中找到 {len(target_elements)} 个svg-gauge-value-and-units元素")
                    
                    # 进一步筛选包含svg-gauge-value-and-units-horizontal class的元素
                    for element in target_elements:
                        try:
                            # 使用更安全的方式获取class属性
                            class_attr = driver.execute_script("return arguments[0].className.baseVal || arguments[0].className;", element)
                            if "svg-gauge-value-and-units-horizontal" in str(class_attr):
                                data_value_svg3 = element.text.strip()
                                logger.info(f"找到SVG3目标元素，class: {class_attr}")
                                logger.info(f"SVG3目标元素文本内容: '{data_value_svg3}'")
                                break
                        except Exception as attr_e:
                            logger.warning(f"获取SVG3元素属性时出错: {str(attr_e)}")
                            continue
                    
                    if data_value_svg3:
                        logger.info(f"成功从SVG3提取到数据: {data_value_svg3}")
                    else:
                        logger.warning("未找到SVG3中包含svg-gauge-value-and-units-horizontal class的元素或元素文本为空")
                        
                except Exception as svg_e:
                    logger.error(f"处理第3个SVG时出错: {str(svg_e)}")
                    logger.error(f"详细错误信息: {traceback.format_exc()}")
            else:
                logger.warning(f"页面中SVG元素数量不足，只找到 {len(svg_elements)} 个，需要至少3个")
            
            # 查找第5个SVG元素
            if len(svg_elements) >= 5:
                svg_element = svg_elements[4]  # 第5个SVG（索引为4）
                logger.info("正在查找第5个SVG中的特定class元素")
                
                try:
                    # 在第5个SVG内查找class为svg-gauge-value-and-units svg-gauge-value-and-units-horizontal的元素
                    target_elements = svg_element.find_elements(By.CLASS_NAME, "svg-gauge-value-and-units")
                    logger.info(f"在第5个SVG中找到 {len(target_elements)} 个svg-gauge-value-and-units元素")
                    
                    # 进一步筛选包含svg-gauge-value-and-units-horizontal class的元素
                    for element in target_elements:
                        try:
                            # 使用更安全的方式获取class属性
                            class_attr = driver.execute_script("return arguments[0].className.baseVal || arguments[0].className;", element)
                            if "svg-gauge-value-and-units-horizontal" in str(class_attr):
                                data_value_svg5 = element.text.strip()
                                logger.info(f"找到SVG5目标元素，class: {class_attr}")
                                logger.info(f"SVG5目标元素文本内容: '{data_value_svg5}'")
                                break
                        except Exception as attr_e:
                            logger.warning(f"获取SVG5元素属性时出错: {str(attr_e)}")
                            continue
                    
                    if data_value_svg5:
                        logger.info(f"成功从SVG5提取到数据: {data_value_svg5}")
                    else:
                        logger.warning("未找到SVG5中包含svg-gauge-value-and-units-horizontal class的元素或元素文本为空")
                        
                except Exception as svg_e:
                    logger.error(f"处理第5个SVG时出错: {str(svg_e)}")
                    logger.error(f"详细错误信息: {traceback.format_exc()}")
            else:
                logger.warning(f"页面中SVG元素数量不足，只找到 {len(svg_elements)} 个，需要至少5个")
                        
        except Exception as e:
            logger.error(f"SVG查找失败: {str(e)}")
            logger.error(f"详细错误信息: {traceback.format_exc()}")
        
        # 提取J列数据（新功能）
        j_column_data = extract_svg_data_for_j_column(driver, page_num)
        
        # 如果没有数据，设置默认值
        if not data_value_svg3:
            logger.warning(f"第{page_num}个页面SVG3未能提取到数据")
            data_value_svg3 = f"页面{page_num}SVG3数据提取失败 - {datetime.now().strftime('%H:%M:%S')}"
        
        if not data_value_svg5:
            logger.warning(f"第{page_num}个页面SVG5未能提取到数据")
            data_value_svg5 = f"页面{page_num}SVG5数据提取失败 - {datetime.now().strftime('%H:%M:%S')}"
        
        # 保存数据到Excel
        logger.info(f"准备将数据保存到Excel单元格 {cell_address_h}, {cell_address_i}, {cell_address_j}")
        excel_file = os.path.join(date_folder, "日常检查表.xlsx")
        if os.path.exists(excel_file):
            try:
                workbook = load_workbook(excel_file)
                worksheet = workbook.active
                worksheet[cell_address_h] = data_value_svg3
                worksheet[cell_address_i] = data_value_svg5
                worksheet[cell_address_j] = j_column_data  # 新功能：保存J列数据
                workbook.save(excel_file)
                logger.info(f"数据已成功保存到 {cell_address_h} 单元格: {data_value_svg3}")
                logger.info(f"数据已成功保存到 {cell_address_i} 单元格: {data_value_svg5}")
                logger.info(f"数据已成功保存到 {cell_address_j} 单元格: {j_column_data}")
            except Exception as excel_e:
                logger.error(f"保存到Excel时出错: {str(excel_e)}")
                logger.error(f"详细错误信息: {traceback.format_exc()}")
        else:
            logger.error(f"Excel文件 {excel_file} 不存在")
            
    except Exception as e:
        logger.error(f"提取数据时发生严重错误: {str(e)}")
        logger.error(f"详细错误信息: {traceback.format_exc()}")
        
        # 即使出错也要保存错误信息到Excel
        try:
            excel_file = os.path.join(date_folder, "日常检查表.xlsx")
            if os.path.exists(excel_file):
                workbook = load_workbook(excel_file)
                worksheet = workbook.active
                worksheet[cell_address_h] = f"提取失败: {str(e)[:50]}"
                worksheet[cell_address_i] = f"提取失败: {str(e)[:50]}"
                worksheet[cell_address_j] = f"C:\\提取失败\nD:\\提取失败"
                workbook.save(excel_file)
                logger.info(f"错误信息已保存到Excel单元格 {cell_address_h}, {cell_address_i}, {cell_address_j}")
        except Exception as excel_error:
            logger.error(f"保存错误信息到Excel失败: {str(excel_error)}")

def take_screenshots():
    logger.info("开始执行网页截图和数据提取任务")
    
    # 获取当前日期
    today = datetime.now().strftime("%Y%m%d")
    logger.info(f"当前日期: {today}")
    
    # 创建日期文件夹
    date_folder = create_date_folder(today)
    logger.info(f"图片将保存到文件夹: {date_folder}")
    
    # 网页URL列表
    urls = [
        "http://192.168.166.108/zabbix.php?action=dashboard.view&dashboardid=392&page=2",
        "http://192.168.166.108/zabbix.php?action=dashboard.view&dashboardid=392&page=3",
        "http://192.168.166.108/zabbix.php?action=dashboard.view&dashboardid=392&page=4",
        "http://192.168.166.108/zabbix.php?action=dashboard.view&dashboardid=392&page=5"
    ]
    
    # 截图文件名列表（包含文件夹路径）
    filenames = [
        os.path.join(date_folder, f"{today}_WMS1.PNG"),
        os.path.join(date_folder, f"{today}_WMS2.PNG"),
        os.path.join(date_folder, f"{today}_QZPMS.PNG"),
        os.path.join(date_folder, f"{today}_QNPMS.PNG")
    ]
    
    # Excel单元格地址列表
    excel_cells_h = ["H4", "H5", "H6", "H7"]
    excel_cells_i = ["I4", "I5", "I6", "I7"]
    excel_cells_j = ["J4", "J5", "J6", "J7"]  # 新增J列单元格
    
    # 设置Edge浏览器
    edge_driver_path = os.path.join(os.getcwd(), "msedgedriver.exe")
    logger.info(f"Edge驱动路径: {edge_driver_path}")
    
    if not os.path.exists(edge_driver_path):
        logger.error(f"Edge驱动文件不存在: {edge_driver_path}")
        return
    
    service = Service(edge_driver_path)
    
    # 设置浏览器选项
    options = webdriver.EdgeOptions()
    options.add_argument("--headless")  # 无头模式
    options.add_argument("--disable-gpu")
    options.add_argument("--no-sandbox")
    options.add_argument("--window-size=1920,1080")
    options.add_argument("--disable-dev-shm-usage")
    options.add_argument("--disable-extensions")
    options.add_argument("--disable-plugins")
    # 添加pyinstaller兼容性选项
    options.add_argument("--disable-web-security")
    options.add_argument("--allow-running-insecure-content")
    
    logger.info("浏览器选项配置完成")
    
    try:
        # 启动浏览器
        logger.info("正在启动Edge浏览器...")
        driver = webdriver.Edge(service=service, options=options)
        logger.info("Edge浏览器启动成功")
        
        wait = WebDriverWait(driver, 15)
        
        # 先访问登录页面
        login_url = "http://192.168.166.108"
        logger.info(f"访问登录页面: {login_url}")
        driver.get(login_url)
        
        # 等待登录页面加载
        time.sleep(2)
        logger.info("登录页面加载完成")
        
        try:
            # 查找用户名输入框
            logger.info("查找用户名输入框...")
            username_field = wait.until(EC.presence_of_element_located((By.NAME, "name")))
            username_field.clear()
            username_field.send_keys("guest")
            logger.info("用户名输入完成")
            
            # 查找密码输入框
            logger.info("查找密码输入框...")
            password_field = driver.find_element(By.NAME, "password")
            password_field.clear()
            password_field.send_keys("")  # guest账户密码通常为空
            logger.info("密码输入完成")
            
            # 点击登录按钮
            logger.info("点击登录按钮...")
            login_button = driver.find_element(By.NAME, "enter")
            login_button.click()
            
            # 等待登录完成
            time.sleep(2)
            logger.info("登录流程完成")
            
        except Exception as login_error:
            logger.error(f"登录过程中出现错误: {str(login_error)}")
            logger.error(f"详细错误信息: {traceback.format_exc()}")
            logger.info("尝试直接访问页面...")
        
        # 依次访问每个网页，先提取数据再截图
        for i, (url, filename, cell_h, cell_i, cell_j) in enumerate(zip(urls, filenames, excel_cells_h, excel_cells_i, excel_cells_j)):
            logger.info(f"\n{'='*50}")
            logger.info(f"开始处理第{i+1}个网页")
            logger.info(f"URL: {url}")
            logger.info(f"文件名: {filename}")
            logger.info(f"Excel单元格: {cell_h}, {cell_i}, {cell_j}")
            logger.info(f"{'='*50}")
            
            try:
                # 访问网页
                logger.info(f"正在访问网页...")
                driver.get(url)
                
                # 等待页面加载完成
                time.sleep(2)
                logger.info("页面访问完成")
                
                # 先提取数据到Excel（写入日期文件夹中的副本）
                logger.info("开始数据提取...")
                extract_data_to_excel(driver, cell_h, cell_i, cell_j, i+1, date_folder)
                
                # 再进行截图
                logger.info("开始截图...")
                
                # 先截取整个页面到临时文件
                temp_filename = f"temp_{i}.png"
                driver.save_screenshot(temp_filename)
                logger.info(f"临时截图保存: {temp_filename}")
                
                # 使用PIL裁剪指定区域
                img = Image.open(temp_filename)
                cropped = img.crop((180, 120, 1850, 870))
                cropped.save(filename)
                logger.info(f"裁剪后截图保存: {filename}")
                
                # 删除临时文件
                os.remove(temp_filename)
                logger.info("临时文件已删除")
                
                # 等待一下再访问下一个网页
                time.sleep(2)
                
            except Exception as page_error:
                logger.error(f"处理第{i+1}个网页时出错: {str(page_error)}")
                logger.error(f"详细错误信息: {traceback.format_exc()}")
                continue
        
        logger.info("\n所有网页处理完成！")
        
        # 合并四张图片
        logger.info("开始合并图片...")
        combine_images(filenames, today, date_folder)
        
    except Exception as e:
        logger.error(f"程序执行过程中发生严重错误: {str(e)}")
        logger.error(f"详细错误信息: {traceback.format_exc()}")
    
    finally:
        # 关闭浏览器
        if 'driver' in locals():
            logger.info("正在关闭浏览器...")
            driver.quit()
            logger.info("浏览器已关闭")

def combine_images(filenames, today, date_folder):
    """将四张图片合并为一张，并添加浅绿色分隔线"""
    logger.info("开始合并图片")
    
    try:
        # 打开四张图片
        images = []
        for filename in filenames:
            if os.path.exists(filename):
                img = Image.open(filename)
                images.append(img)
                logger.info(f"成功加载图片: {filename}")
            else:
                logger.error(f"图片文件不存在: {filename}")
                return
        
        if len(images) != 4:
            logger.error(f"图片数量不正确，需要4张，实际{len(images)}张")
            return
        
        # 获取图片尺寸
        img_width, img_height = images[0].size
        logger.info(f"图片尺寸: {img_width} x {img_height}")
        
        # 分隔线宽度
        separator_width = 10
        
        # 创建新的画布
        combined_width = img_width * 2 + separator_width
        combined_height = img_height * 2 + separator_width
        combined_image = Image.new('RGB', (combined_width, combined_height), 'white')
        logger.info(f"合并画布尺寸: {combined_width} x {combined_height}")
        
        # 创建绘图对象
        draw = ImageDraw.Draw(combined_image)
        
        # 浅绿色
        light_green = (144, 238, 144)
        
        # 粘贴图片到指定位置
        combined_image.paste(images[0], (0, 0))  # 左上
        combined_image.paste(images[1], (img_width + separator_width, 0))  # 右上
        combined_image.paste(images[2], (0, img_height + separator_width))  # 左下
        combined_image.paste(images[3], (img_width + separator_width, img_height + separator_width))  # 右下
        
        # 绘制分隔线
        draw.rectangle([
            (img_width, 0),
            (img_width + separator_width, combined_height)
        ], fill=light_green)
        
        draw.rectangle([
            (0, img_height),
            (combined_width, img_height + separator_width)
        ], fill=light_green)
        
        # 保存合并后的图片到日期文件夹
        combined_filename = os.path.join(date_folder, f"{today}-机房.png")
        combined_image.save(combined_filename)
        logger.info(f"图片合并完成，已保存为: {combined_filename}")
        
    except Exception as e:
        logger.error(f"合并图片时发生错误: {str(e)}")
        logger.error(f"详细错误信息: {traceback.format_exc()}")

if __name__ == "__main__":
    logger.info("程序开始执行")
    take_screenshots()
    logger.info("程序执行结束")